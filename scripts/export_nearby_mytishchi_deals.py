#!/usr/bin/env python3
import csv
import sqlite3
import statistics
from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
DB = ROOT / "Сделки Росреестр" / "output" / "rosreestr_deals_unified_2025q3_2026q1.sqlite"
VISIBLE_DIR = ROOT / "Сделки Росреестр" / "Ближние деревни"
OUTPUT_DIR = ROOT / "Сделки Росреестр" / "output"

DETAIL_NAME = "сделки_ближние_деревни_мытищи_2025q3_2026q1.csv"
SUMMARY_NAME = "сводка_ближние_деревни_мытищи_2025q3_2026q1.csv"
XLSX_NAME = "сделки_ближние_деревни_мытищи_2025q3_2026q1.xlsx"


QUERY = """
WITH local AS (
  SELECT
    CASE
      WHEN city_name = 'Осташково' THEN 'Осташково (проверка вместо Осташкино)'
      WHEN city_name = 'Жостово' THEN 'Жостово (в запросе: Жестово)'
      WHEN city_name = 'Никульское' THEN 'Никульское'
      WHEN city_name = 'Сорокино' THEN 'Сорокино'
      WHEN city_name = 'Манюхино' THEN 'Манюхино'
      WHEN city_name = 'Витенево' THEN 'Витенево'
      WHEN city_name = 'Чиверево' THEN 'Чиверево'
      WHEN city_name = 'Пирогово' AND street_name = 'Пирогово' THEN 'СНТ Пирогово'
    END AS area_name,
    *
  FROM v_deals
  WHERE region_code = 50
    AND district_name = 'Мытищи'
    AND (
      city_name IN (
        'Осташково', 'Жостово', 'Никульское',
        'Сорокино', 'Манюхино', 'Витенево', 'Чиверево'
      )
      OR (city_name = 'Пирогово' AND street_name = 'Пирогово')
    )
)
SELECT
  area_name,
  source_year,
  source_quarter,
  period_start_date,
  district_name,
  city_name,
  street_name,
  district,
  city,
  quarter_cad_number,
  street,
  realestate_type_name,
  doc_type,
  number,
  area,
  deal_price,
  CASE
    WHEN realestate_type_name = 'Земельный участок' AND area > 0
    THEN deal_price / area * 100
  END AS price_per_sotka,
  CASE
    WHEN realestate_type_name != 'Земельный участок' AND area > 0
    THEN deal_price / area
  END AS price_per_m2,
  purpose_code,
  wall_material_code,
  year_build,
  floor,
  source_file
FROM local
ORDER BY area_name, period_start_date, realestate_type_name, deal_price;
"""


DETAIL_HEADERS = [
    "Локация",
    "Год",
    "Квартал",
    "Дата периода",
    "Район",
    "Населенный пункт",
    "Улица",
    "Район в Росреестре",
    "Населенный пункт в Росреестре",
    "Кадастровый квартал",
    "Улица в Росреестре",
    "Тип объекта",
    "Тип документа",
    "Количество сделок в строке",
    "Площадь",
    "Цена сделки",
    "Цена за сотку",
    "Цена за м2",
    "Код назначения",
    "Код материала стен",
    "Год постройки",
    "Этаж",
    "Источник",
]

DETAIL_KEYS = [
    "area_name",
    "source_year",
    "source_quarter",
    "period_start_date",
    "district_name",
    "city_name",
    "street_name",
    "district",
    "city",
    "quarter_cad_number",
    "street",
    "realestate_type_name",
    "doc_type",
    "number",
    "area",
    "deal_price",
    "price_per_sotka",
    "price_per_m2",
    "purpose_code",
    "wall_material_code",
    "year_build",
    "floor",
    "source_file",
]

SUMMARY_HEADERS = [
    "Локация",
    "Тип объекта",
    "Строк в датасете",
    "Количество сделок",
    "Средняя цена",
    "Медиана цены",
    "Минимальная цена",
    "Максимальная цена",
    "Средняя площадь",
    "Средняя удельная цена",
    "Медиана удельной цены",
    "Единица удельной цены",
]


def weighted_values(rows, value_key):
    values = []
    for row in rows:
        value = row.get(value_key)
        if value is None:
            continue
        values.extend([float(value)] * int(row.get("number") or 1))
    return values


def build_summary(rows):
    groups = defaultdict(list)
    for row in rows:
        groups[(row["area_name"], row["realestate_type_name"])].append(row)

    summary = []
    for (area_name, object_type), group in sorted(groups.items()):
        total_count = sum(int(row.get("number") or 1) for row in group)
        total_price = sum((float(row["deal_price"] or 0) * int(row.get("number") or 1)) for row in group)
        total_area = sum((float(row["area"] or 0) * int(row.get("number") or 1)) for row in group)
        price_values = weighted_values(group, "deal_price")
        unit_key = "price_per_sotka" if object_type == "Земельный участок" else "price_per_m2"
        unit_values = weighted_values(group, unit_key)

        summary.append(
            {
                "Локация": area_name,
                "Тип объекта": object_type,
                "Строк в датасете": len(group),
                "Количество сделок": total_count,
                "Средняя цена": round(total_price / total_count) if total_count else None,
                "Медиана цены": round(statistics.median(price_values)) if price_values else None,
                "Минимальная цена": round(min(price_values), 2) if price_values else None,
                "Максимальная цена": round(max(price_values), 2) if price_values else None,
                "Средняя площадь": round(total_area / total_count, 1) if total_count else None,
                "Средняя удельная цена": round(statistics.fmean(unit_values)) if unit_values else None,
                "Медиана удельной цены": round(statistics.median(unit_values)) if unit_values else None,
                "Единица удельной цены": "руб/сотка" if object_type == "Земельный участок" else "руб/м2",
            }
        )
    return summary


def write_csv(path, headers, rows):
    with path.open("w", encoding="utf-8-sig", newline="") as file:
        writer = csv.DictWriter(file, fieldnames=headers, delimiter=";")
        writer.writeheader()
        writer.writerows(rows)


def write_workbook(path, summary_rows, detail_rows):
    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "Сводка"
    ws_detail = wb.create_sheet("Детали")

    for ws, headers, rows in (
        (ws_summary, SUMMARY_HEADERS, summary_rows),
        (ws_detail, DETAIL_HEADERS, detail_rows),
    ):
        ws.append(headers)
        for row in rows:
            ws.append([row.get(header) for header in headers])

        header_fill = PatternFill("solid", fgColor="D9EAF7")
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(wrap_text=True, vertical="top")

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for col_idx, header in enumerate(headers, 1):
            max_len = len(str(header))
            for row in ws.iter_rows(min_col=col_idx, max_col=col_idx, min_row=2, max_row=ws.max_row):
                value = row[0].value
                if value is not None:
                    max_len = max(max_len, len(str(value)))
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 10), 36)

    wb.save(path)


def export_to(folder):
    folder.mkdir(parents=True, exist_ok=True)
    conn = sqlite3.connect(DB)
    conn.row_factory = sqlite3.Row
    try:
        raw_rows = [dict(row) for row in conn.execute(QUERY)]
    finally:
        conn.close()

    detail_rows = [
        {header: row[key] for header, key in zip(DETAIL_HEADERS, DETAIL_KEYS)}
        for row in raw_rows
    ]
    summary_rows = build_summary(raw_rows)

    write_csv(folder / DETAIL_NAME, DETAIL_HEADERS, detail_rows)
    write_csv(folder / SUMMARY_NAME, SUMMARY_HEADERS, summary_rows)
    write_workbook(folder / XLSX_NAME, summary_rows, detail_rows)
    return folder / DETAIL_NAME, folder / SUMMARY_NAME, folder / XLSX_NAME, len(detail_rows), len(summary_rows)


def main():
    results = []
    for folder in (VISIBLE_DIR, OUTPUT_DIR):
        results.append(export_to(folder))
    for detail, summary, xlsx, detail_count, summary_count in results:
        print(f"folder={detail.parent}")
        print(f"detail={detail} rows={detail_count}")
        print(f"summary={summary} rows={summary_count}")
        print(f"xlsx={xlsx}")


if __name__ == "__main__":
    main()
