#!/usr/bin/env python3
import math
import sqlite3
import statistics
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
DB = ROOT / "Сделки Росреестр" / "output" / "rosreestr_deals_unified_2025q3_2026q1.sqlite"
OUT_DIR = ROOT / "Сделки Росреестр" / "Кварталы МОНАВИ Агаларов Миллениум"
OUT_XLSX = OUT_DIR / "единый_пакетный_анализ_участки_дома_премиум_2025q3_2026q1.xlsx"

QUARTER_META = {
    "50:11:0050115": ("КП МОНАВИ / дер. Маслово", "МОНАВИ"),
    "50:08:0050432": ("КП Агаларов Эстейт / КП Миллениум Парк", "Агаларов/Миллениум"),
    "50:08:0050415": ("КП Агаларов Эстейт / КП Миллениум Парк", "Агаларов/Миллениум"),
    "50:08:0050421": ("КП Агаларов Эстейт / КП Миллениум Парк", "Агаларов/Миллениум"),
    "50:08:0050429": ("КП Агаларов Эстейт / КП Миллениум Парк", "Агаларов/Миллениум"),
    "50:08:0050435": ("КП Агаларов Эстейт / КП Миллениум Парк", "Агаларов/Миллениум"),
    "50:08:0050401": ("КП Агаларов Эстейт / КП Миллениум Парк", "Агаларов/Миллениум"),
    "50:08:0050407": ("КП Агаларов Эстейт / КП Миллениум Парк", "Агаларов/Миллениум"),
    "50:08:0050426": ("КП Агаларов Эстейт / КП Миллениум Парк", "Агаларов/Миллениум"),
}


def quantile(values, p):
    values = sorted(values)
    if not values:
        return None
    k = (len(values) - 1) * p
    floor = math.floor(k)
    ceil = math.ceil(k)
    if floor == ceil:
        return values[floor]
    return values[floor] * (ceil - k) + values[ceil] * (k - floor)


def money(value):
    if value is None:
        return None
    return round(float(value))


def stats(values):
    values = [float(v) for v in values if v is not None]
    if not values:
        return {
            "n": 0,
            "avg": None,
            "median": None,
            "p25": None,
            "p75": None,
            "min": None,
            "max": None,
        }
    return {
        "n": len(values),
        "avg": money(statistics.fmean(values)),
        "median": money(statistics.median(values)),
        "p25": money(quantile(values, 0.25)),
        "p75": money(quantile(values, 0.75)),
        "min": round(min(values), 2),
        "max": round(max(values), 2),
    }


def object_type(code):
    return {
        "002001001000": "Земельный участок",
        "002001002000": "Здание",
        "002001003000": "Помещение",
        "002001009000": "Машиноместо",
    }.get(code, code)


def fetch_rows():
    placeholders = ",".join("?" for _ in QUARTER_META)
    conn = sqlite3.connect(DB)
    conn.row_factory = sqlite3.Row
    try:
        rows = [
            dict(row)
            for row in conn.execute(
                f"""
                SELECT
                    id, number, okato, region_code, district, city, quarter_cad_number,
                    street, realestate_type_code, wall_material_code, year_build, floor,
                    purpose_code, area, period_start_date, deal_price, currency, doc_type,
                    source_file, source_year, source_quarter
                FROM deals
                WHERE quarter_cad_number IN ({placeholders})
                  AND deal_price IS NOT NULL
                  AND deal_price > 0
                ORDER BY quarter_cad_number, period_start_date, city, street, deal_price
                """,
                tuple(QUARTER_META.keys()),
            )
        ]
    finally:
        conn.close()

    for row in rows:
        label, group = QUARTER_META[row["quarter_cad_number"]]
        row["project_label"] = label
        row["project_group"] = group
        row["object_type"] = object_type(row["realestate_type_code"])
    return rows


def package_key(row):
    # Best available proxy: no cadastral numbers are published, so identical price/date/location
    # is the strongest signal that rows belong to one economic package.
    return (
        row["quarter_cad_number"],
        row["period_start_date"],
        row.get("city") or "",
        row.get("street") or "",
        round(float(row["deal_price"]), 2),
        row.get("doc_type") or "",
    )


def package_confidence(group):
    types = {row["object_type"] for row in group}
    has_land = "Земельный участок" in types
    has_building = "Здание" in types
    has_street = any((row.get("street") or "").strip() for row in group)
    if has_land and has_building and has_street:
        return "Высокая: дом+земля с той же ценой и улицей"
    if has_land and has_building:
        return "Средняя: дом+земля с той же ценой, но без улицы"
    if has_land and len(group) > 1:
        return "Средняя: земельный пул с той же ценой"
    if len(group) > 1:
        return "Средняя: несколько объектов с той же ценой"
    return "Одиночная строка"


def build_packages(rows):
    grouped = defaultdict(list)
    for row in rows:
        grouped[package_key(row)].append(row)

    packages = []
    object_rows = []
    for idx, (key, group) in enumerate(
        sorted(grouped.items(), key=lambda item: (item[0][0], item[0][1], item[0][2], item[0][3], item[0][4])),
        start=1,
    ):
        quarter, period, city, street, price, doc_type = key
        label, project_group = QUARTER_META[quarter]
        types = Counter(row["object_type"] for row in group)
        numbers = [int(row["number"] or 1) for row in group]
        estimated_deals = max(numbers) if numbers else 1
        land_rows = [row for row in group if row["object_type"] == "Земельный участок"]
        building_rows = [row for row in group if row["object_type"] == "Здание"]
        premise_rows = [row for row in group if row["object_type"] == "Помещение"]
        land_area_m2 = sum(float(row["area"] or 0) for row in land_rows)
        building_area_m2 = sum(float(row["area"] or 0) for row in building_rows)
        premise_area_m2 = sum(float(row["area"] or 0) for row in premise_rows)
        land_area_sotka = land_area_m2 / 100 if land_area_m2 else None
        package_price_per_sotka = price / land_area_m2 * 100 if land_area_m2 else None
        package_price_per_building_m2 = price / building_area_m2 if building_area_m2 else None
        package_type = (
            "Дом + земля"
            if land_rows and building_rows
            else "Земельный пакет"
            if land_rows
            else "Дом/здание без видимой земли"
            if building_rows
            else "Помещения/прочее"
        )
        districts = ", ".join(sorted({row["district"] for row in group if row.get("district")}))
        sources = ", ".join(sorted({row["source_file"] for row in group if row.get("source_file")}))
        years = [row["year_build"] for row in building_rows if row.get("year_build") is not None]

        package = {
            "package_id": idx,
            "Группа": project_group,
            "Проект/зона": label,
            "Кадастровый квартал": quarter,
            "Период": period,
            "Район": districts,
            "Населенный пункт": city,
            "Улица": street,
            "Тип пакета": package_type,
            "Надежность склейки": package_confidence(group),
            "Цена пакета": price,
            "Оценка кол-ва сделок": estimated_deals,
            "Объем пакета": price * estimated_deals,
            "Строк объектов": len(group),
            "Состав": "; ".join(f"{typ}: {count}" for typ, count in sorted(types.items())),
            "Земля строк": len(land_rows),
            "Здания строк": len(building_rows),
            "Помещения строк": len(premise_rows),
            "Площадь земли м2": land_area_m2 or None,
            "Площадь земли соток": round(land_area_sotka, 2) if land_area_sotka else None,
            "Площадь зданий м2": building_area_m2 or None,
            "Площадь помещений м2": premise_area_m2 or None,
            "Цена пакета / сотка": money(package_price_per_sotka),
            "Цена пакета / м2 здания": money(package_price_per_building_m2),
            "Годы зданий": ", ".join(str(year) for year in sorted(set(years))) if years else "",
            "Тип документа": doc_type,
            "Источник": sources,
        }
        packages.append(package)

        for row in group:
            object_rows.append(
                {
                    "package_id": idx,
                    "Группа": project_group,
                    "Кадастровый квартал": quarter,
                    "Период": period,
                    "Район": row.get("district"),
                    "Населенный пункт": row.get("city"),
                    "Улица": row.get("street"),
                    "Тип пакета": package_type,
                    "Тип объекта": row["object_type"],
                    "number": row.get("number"),
                    "Площадь": row.get("area"),
                    "Цена строки": row.get("deal_price"),
                    "Год постройки": row.get("year_build"),
                    "Этаж": row.get("floor"),
                    "Код назначения/ВРИ": row.get("purpose_code"),
                    "Код материала": row.get("wall_material_code"),
                    "Источник": row.get("source_file"),
                }
            )

    return packages, object_rows


def expand_by_deals(packages, value_key):
    values = []
    for package in packages:
        value = package.get(value_key)
        if value is None:
            continue
        values.extend([float(value)] * int(package["Оценка кол-ва сделок"] or 1))
    return values


def make_summary(packages):
    rows = []
    for group_name in sorted({p["Группа"] for p in packages}):
        group_packages = [p for p in packages if p["Группа"] == group_name]
        for package_type in sorted({p["Тип пакета"] for p in group_packages}):
            subset = [p for p in group_packages if p["Тип пакета"] == package_type]
            prices = expand_by_deals(subset, "Цена пакета")
            per_sotka = expand_by_deals(subset, "Цена пакета / сотка")
            per_m2 = expand_by_deals(subset, "Цена пакета / м2 здания")
            price_stats = stats(prices)
            sotka_stats = stats(per_sotka)
            m2_stats = stats(per_m2)
            rows.append(
                {
                    "Группа": group_name,
                    "Тип пакета": package_type,
                    "Пакетов": len(subset),
                    "Оценка кол-ва сделок": sum(int(p["Оценка кол-ва сделок"] or 1) for p in subset),
                    "Объем, руб": money(sum(float(p["Объем пакета"]) for p in subset)),
                    "Средняя цена пакета": price_stats["avg"],
                    "Медиана цены пакета": price_stats["median"],
                    "P25 цена пакета": price_stats["p25"],
                    "P75 цена пакета": price_stats["p75"],
                    "Средняя руб/сотка": sotka_stats["avg"],
                    "Медиана руб/сотка": sotka_stats["median"],
                    "P25 руб/сотка": sotka_stats["p25"],
                    "P75 руб/сотка": sotka_stats["p75"],
                    "Средняя руб/м2 здания": m2_stats["avg"],
                    "Медиана руб/м2 здания": m2_stats["median"],
                    "Пакетов >= 1 млрд": len([p for p in subset if float(p["Цена пакета"]) >= 1_000_000_000]),
                    "Сделок >= 1 млрд": sum(
                        int(p["Оценка кол-ва сделок"] or 1)
                        for p in subset
                        if float(p["Цена пакета"]) >= 1_000_000_000
                    ),
                }
            )
    return rows


def make_quarter_summary(packages):
    rows = []
    for quarter in QUARTER_META:
        subset = [p for p in packages if p["Кадастровый квартал"] == quarter]
        if not subset:
            label, group = QUARTER_META[quarter]
            rows.append(
                {
                    "Группа": group,
                    "Проект/зона": label,
                    "Кадастровый квартал": quarter,
                    "Пакетов": 0,
                    "Оценка кол-ва сделок": 0,
                    "Объем, руб": 0,
                    "Медиана цены пакета": None,
                    "Медиана руб/сотка": None,
                    "Медиана руб/м2 здания": None,
                    "Пакетов >= 1 млрд": 0,
                }
            )
            continue
        label, group = QUARTER_META[quarter]
        prices = expand_by_deals(subset, "Цена пакета")
        per_sotka = expand_by_deals(subset, "Цена пакета / сотка")
        per_m2 = expand_by_deals(subset, "Цена пакета / м2 здания")
        rows.append(
            {
                "Группа": group,
                "Проект/зона": label,
                "Кадастровый квартал": quarter,
                "Пакетов": len(subset),
                "Оценка кол-ва сделок": sum(int(p["Оценка кол-ва сделок"] or 1) for p in subset),
                "Объем, руб": money(sum(float(p["Объем пакета"]) for p in subset)),
                "Медиана цены пакета": stats(prices)["median"],
                "Медиана руб/сотка": stats(per_sotka)["median"],
                "Медиана руб/м2 здания": stats(per_m2)["median"],
                "Пакетов >= 1 млрд": len([p for p in subset if float(p["Цена пакета"]) >= 1_000_000_000]),
            }
        )
    return rows


def make_conclusions(packages, summary_rows):
    all_packages = len(packages)
    all_deals = sum(int(p["Оценка кол-ва сделок"] or 1) for p in packages)
    ge_1b = [p for p in packages if float(p["Цена пакета"]) >= 1_000_000_000]
    house_land = [p for p in packages if p["Тип пакета"] == "Дом + земля"]
    land_only = [p for p in packages if p["Тип пакета"] == "Земельный пакет"]
    top = sorted(packages, key=lambda p: float(p["Цена пакета"]), reverse=True)[:5]
    top_text = "; ".join(
        f"{p['Кадастровый квартал']} {p['Населенный пункт']} {p['Улица'] or ''}: {money(p['Цена пакета']):,}".replace(",", " ")
        for p in top
    )

    return [
        {
            "Вывод": "Объектные строки склеены в экономические пакеты",
            "Комментарий": f"Из {sum(p['Строк объектов'] for p in packages)} объектных строк получилось {all_packages} пакетов; оценка количества сделок по пакетам: {all_deals}.",
        },
        {
            "Вывод": "Миллиардный хвост сохраняется и после склейки",
            "Комментарий": f"Пакетов >= 1 млрд: {len(ge_1b)}. Это {round(100 * len(ge_1b) / all_packages, 2)}% от пакетов. По оценке количества сделок: {sum(int(p['Оценка кол-ва сделок'] or 1) for p in ge_1b)}.",
        },
        {
            "Вывод": "Самый дорогой подтвержденный дом+земля",
            "Комментарий": "50:08:0050435, Захарово, Правобережная: пакет 2.015 млрд, земля 36.46 сотки, здания 2008.3 м2 + 52.1 м2.",
        },
        {
            "Вывод": "Часть миллиардных строк - не дома, а земельные пулы/земельные строки",
            "Комментарий": "Например 50:08:0050435, Захарово, Заречная: земля 19673 м2, number=5, цена 1.054 млрд; это лучше трактовать как земельный пул, а не как один обычный участок.",
        },
        {
            "Вывод": "Дома без видимой земли в этих кварталах тоже есть",
            "Комментарий": "Если рядом нет земельной строки с той же ценой, дом нельзя корректно перевести в цену сотки; его надо анализировать как цену здания/пакета по м2 дома.",
        },
        {
            "Вывод": "Топ-5 цен пакетов",
            "Комментарий": top_text,
        },
        {
            "Вывод": "Для premium-аналитики лучше использовать пакетную таблицу",
            "Комментарий": "Объектные строки полезны для состава, но для объема рынка и доли сделок >=1 млрд лучше использовать лист 'Пакеты', иначе дом+земля могут задваивать цену договора.",
        },
    ]


def write_workbook(packages, object_rows, summary_rows, quarter_rows, conclusions):
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    wb = Workbook()

    def add_sheet(title, data):
        ws = wb.active if wb.active.max_row == 1 and wb.active["A1"].value is None else wb.create_sheet(title[:31])
        ws.title = title[:31]
        headers = list(data[0].keys()) if data else ["Нет данных"]
        ws.append(headers)
        for row in data:
            ws.append([row.get(header) for header in headers])
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="D9EAF7")
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for col_idx in range(1, len(headers) + 1):
            max_len = max(
                len(str(ws.cell(row_idx, col_idx).value or ""))
                for row_idx in range(1, min(ws.max_row, 300) + 1)
            )
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 10), 46)
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                if isinstance(cell.value, (int, float)):
                    cell.number_format = "#,##0"

    method = [
        {
            "Пункт": "Склейка в пакет",
            "Описание": "Группировка по кадастровому кварталу, периоду, населенному пункту, улице, цене и типу документа. Это лучший доступный прокси, потому что кадастровые номера в открытом датасете скрыты.",
        },
        {
            "Пункт": "Цена пакета",
            "Описание": "Внутри пакета цена берется один раз, даже если она повторяется у земли и здания. Строки объектов не суммируются.",
        },
        {
            "Пункт": "Оценка количества сделок",
            "Описание": "Для пакета используется максимум поля number среди строк пакета. Так пакет дом+земля с number=1 не превращается в две сделки.",
        },
        {
            "Пункт": "Цена сотки",
            "Описание": "Для пакетов с землей: цена пакета / сумма площади земельных строк * 100.",
        },
        {
            "Пункт": "Ограничение",
            "Описание": "Если улица пустая или несколько земельных строк имеют одинаковую цену, склейка вероятная, но не доказанная. Для юридической проверки нужен кадастровый номер, которого в датасете нет.",
        },
    ]

    add_sheet("Выводы", conclusions)
    add_sheet("Сводка пакетов", summary_rows)
    add_sheet("Сводка кварталов", quarter_rows)
    add_sheet("Пакеты", packages)
    add_sheet("Объектные строки", object_rows)
    add_sheet("Методика", method)
    wb.save(OUT_XLSX)


def main():
    rows = fetch_rows()
    packages, object_rows = build_packages(rows)
    summary_rows = make_summary(packages)
    quarter_rows = make_quarter_summary(packages)
    conclusions = make_conclusions(packages, summary_rows)
    write_workbook(packages, object_rows, summary_rows, quarter_rows, conclusions)
    print(OUT_XLSX)
    print(f"object_rows={len(object_rows)} packages={len(packages)}")
    print(f"ge_1b_packages={len([p for p in packages if float(p['Цена пакета']) >= 1_000_000_000])}")


if __name__ == "__main__":
    main()
